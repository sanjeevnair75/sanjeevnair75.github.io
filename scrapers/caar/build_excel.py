"""
CAAR Excel Workbook Builder
============================
Takes caar_all_enriched.json and builds the multi-sheet Excel workbook.

Output: caar_analytics.xlsx
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter


GOLD = "B8974A"
DARK = "0E0E0E"
BG_LIGHT = "FAF8F3"
WHITE = "FFFFFF"

COLUMNS = ['section', 'page_no', 'sl_no_on_page', 'applicant', 'date_application',
           'date_ruling', 'year', 'subject', 'issue_type', 'issue_confidence',
           'ruling_no', 'pdf_link', 'pdf_available']

COLUMN_HEADERS = {
    'section': 'Section', 'page_no': 'Pg', 'sl_no_on_page': 'Sl',
    'applicant': 'Applicant', 'date_application': 'Date of Application',
    'date_ruling': 'Date of Ruling', 'year': 'Year', 'subject': 'Subject',
    'issue_type': 'Issue Type', 'issue_confidence': 'Confidence',
    'ruling_no': 'Ruling No.', 'pdf_link': 'PDF Link', 'pdf_available': 'PDF?',
}

COLUMN_WIDTHS = {
    'section': 17, 'page_no': 4, 'sl_no_on_page': 4, 'applicant': 35,
    'date_application': 13, 'date_ruling': 13, 'year': 7, 'subject': 50,
    'issue_type': 16, 'issue_confidence': 11, 'ruling_no': 28,
    'pdf_link': 35, 'pdf_available': 7,
}


def transform_pdf_url(url):
    """Convert original 404-ing URL to the working API URL."""
    if url and '/CONTENTREPO/' in url:
        return url.replace('/CONTENTREPO/', '/content/pdf/CONTENTREPO/')
    return url


def write_data_sheet(ws, data_rows, sheet_label):
    header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    header_fill = PatternFill('solid', start_color=GOLD)
    header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_border = Border(bottom=Side(style='medium', color=DARK))

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    cell_font = Font(name='Calibri', size=10)

    title_font = Font(name='Calibri', size=18, bold=True, color=DARK)
    subtitle_font = Font(name='Calibri', size=10, italic=True, color="666666")

    ws['A1'] = f'CAAR Litigation Analytics  -  {sheet_label}'
    ws['A1'].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    ws['A2'] = f'{len(data_rows)} records'
    ws['A2'].font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    ws.row_dimensions[1].height = 24

    HEADER_ROW = 4
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=COLUMN_HEADERS[col_key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    ws.row_dimensions[HEADER_ROW].height = 32

    for col_idx, col_key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_key]

    for row_idx, record in enumerate(data_rows, start=HEADER_ROW + 1):
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            value = record.get(col_key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment

            if col_key == 'pdf_link' and value:
                # Apply URL transformation so the Excel link actually works
                working_url = transform_pdf_url(value)
                cell.value = working_url
                cell.hyperlink = working_url
                cell.font = Font(name='Calibri', size=10, color='0563C1', underline='single')

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)


def main():
    print("Building CAAR Excel workbook...")

    try:
        with open('caar_all_enriched.json', 'r', encoding='utf-8') as f:
            records = json.load(f)
    except FileNotFoundError:
        print("ERROR: caar_all_enriched.json not found. Run postprocess.py first.")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet('Summary')

    sections = ['Mumbai Rulings', 'Mumbai Orders', 'Delhi Rulings', 'Delhi Orders']

    ws_all = wb.create_sheet('All Records')
    write_data_sheet(ws_all, records, 'All Records (All Benches and Types)')

    for sec in sections:
        sheet_name = sec.replace(' ', '_')[:31]
        ws = wb.create_sheet(sheet_name)
        sec_records = [r for r in records if r.get('section') == sec]
        write_data_sheet(ws, sec_records, sec)

    # ----- Summary sheet -----
    title_font = Font(name='Calibri', size=22, bold=True, color=DARK)
    section_font = Font(name='Calibri', size=12, bold=True, color=GOLD)
    subtitle_font = Font(name='Calibri', size=10, italic=True, color="666666")

    ws_summary['A1'] = 'CAAR Litigation Analytics'
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:F1')

    ws_summary['A2'] = 'Auto-refreshing summary from 953+ records across Mumbai and Delhi CAAR'
    ws_summary['A2'].font = subtitle_font
    ws_summary.merge_cells('A2:F2')

    ws_summary.row_dimensions[1].height = 30

    ws_summary['A4'] = 'OVERVIEW'
    ws_summary['A4'].font = section_font

    row = 5
    ws_summary.cell(row=row, column=1, value='Total Records').font = Font(name='Calibri', size=10, bold=True)
    ws_summary.cell(row=row, column=2, value='=COUNTA(\'All Records\'!A5:A2000)').font = Font(name='Calibri', size=18, bold=True, color=GOLD)
    row += 1

    ws_summary.cell(row=row, column=1, value='With PDF Available').font = Font(name='Calibri', size=10, bold=True)
    ws_summary.cell(row=row, column=2, value='=COUNTIF(\'All Records\'!M5:M2000,"YES")').font = Font(name='Calibri', size=14, bold=True, color=GOLD)
    row += 2

    # Section breakdown
    ws_summary.cell(row=row, column=1, value='BY SECTION').font = section_font
    row += 1
    hdr_row = row

    header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    header_fill = PatternFill('solid', start_color=GOLD)
    header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col_offset, h in enumerate(['Section', 'Count', 'With PDF', 'PDF %'], start=1):
        c = ws_summary.cell(row=row, column=col_offset, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
    row += 1

    for sec in sections:
        sn = sec.replace(' ', '_')[:31]
        ws_summary.cell(row=row, column=1, value=sec)
        ws_summary.cell(row=row, column=2, value=f"=COUNTA('{sn}'!A5:A2000)")
        ws_summary.cell(row=row, column=3, value=f"=COUNTIF('{sn}'!M5:M2000,\"YES\")")
        ws_summary.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)")
        ws_summary.cell(row=row, column=4).number_format = '0.0%'
        row += 1

    ws_summary.cell(row=row, column=1, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    ws_summary.cell(row=row, column=2, value=f"=SUM(B{hdr_row+1}:B{row-1})").font = Font(name='Calibri', size=11, bold=True)
    ws_summary.cell(row=row, column=3, value=f"=SUM(C{hdr_row+1}:C{row-1})").font = Font(name='Calibri', size=11, bold=True)
    ws_summary.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)").font = Font(name='Calibri', size=11, bold=True)
    ws_summary.cell(row=row, column=4).number_format = '0.0%'
    row += 2

    # Issue breakdown
    ws_summary.cell(row=row, column=1, value='BY ISSUE TYPE').font = section_font
    row += 1
    for col_offset, h in enumerate(['Issue Type', 'Count', 'Share'], start=1):
        c = ws_summary.cell(row=row, column=col_offset, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
    row += 1

    total_records = len(records)
    issue_counter = Counter(r['issue_type'] for r in records)
    for issue, count in issue_counter.most_common():
        ws_summary.cell(row=row, column=1, value=issue)
        ws_summary.cell(row=row, column=2, value=f'=COUNTIF(\'All Records\'!I5:I2000,"{issue}")')
        ws_summary.cell(row=row, column=3, value=f'=IFERROR(B{row}/{total_records},0)')
        ws_summary.cell(row=row, column=3).number_format = '0.0%'
        row += 1
    row += 1

    # Year breakdown
    ws_summary.cell(row=row, column=1, value='BY YEAR').font = section_font
    row += 1
    for col_offset, h in enumerate(['Year', 'Total', 'Mumbai', 'Delhi'], start=1):
        c = ws_summary.cell(row=row, column=col_offset, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
    row += 1

    years = sorted(set(r.get('year', '') for r in records if r.get('year', '')))
    for y in years:
        ws_summary.cell(row=row, column=1, value=y)
        ws_summary.cell(row=row, column=2, value=f'=COUNTIF(\'All Records\'!G5:G2000,"{y}")')
        ws_summary.cell(row=row, column=3, value=f'=COUNTIFS(\'All Records\'!G5:G2000,"{y}",\'All Records\'!A5:A2000,"Mumbai*")')
        ws_summary.cell(row=row, column=4, value=f'=COUNTIFS(\'All Records\'!G5:G2000,"{y}",\'All Records\'!A5:A2000,"Delhi*")')
        row += 1

    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 14
    ws_summary.column_dimensions['C'].width = 14
    ws_summary.column_dimensions['D'].width = 12

    # Move Summary first
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_summary)))

    wb.save('caar_analytics.xlsx')
    print("Saved: caar_analytics.xlsx")


if __name__ == "__main__":
    main()
